import openpyxl as xl
wb = xl.load_workbook("transactions.xlsx") #loads xl workbook and loads an workbook object
sheet = wb["Sheet1"]
cell  = sheet['a1']
cell = sheet.cell(1,1)

for row in range(2,sheet.max_row + 1): #start from row 2 to ignore heading and include last row by adding 1
    cell = sheet.cell(row,3)
    corrected_price  = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

wb.save("transaction2.xlsx")